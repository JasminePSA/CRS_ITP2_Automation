import  openpyxl
#from openpyxl import workbook
wk = openpyxl.load_workbook("C:\\Users\ThomasJ1\AppData\Local\Programs\Python\Python310\Scripts\ITP2 TestData.xlsx")
def retrieve_row_sheet(Sheetname):
    sh = wk[Sheetname]
    row = sh.max_row
    col = sh.max_column
    return row
print(retrieve_row_sheet("TestData"))
